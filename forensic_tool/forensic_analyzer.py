"""
MÓDULO DE ANÁLISE FORENSE
Analisador completo de metadados de arquivos
"""

import os
import json
import hashlib
import threading
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
import logging
from typing import Dict, List, Any, Optional, Generator

# Bibliotecas multimídia
try:
    from PIL import Image, ExifTags
    from PyPDF2 import PdfReader
    import docx
    import openpyxl
    import pptx
    import mutagen
    import cv2
    import zipfile
    import magic
except ImportError as e:
    print(f"❌ Biblioteca não instalada: {e}")
    raise

# Configurações
SUPPORTED_EXTENSIONS = {
    'IMAGENS': ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff', '.tif', '.webp'],
    'PDF': ['.pdf'],
    'DOCUMENTOS': ['.docx', '.doc', '.xlsx', '.xls', '.pptx', '.ppt', '.txt', '.rtf'],
    'AUDIO': ['.mp3', '.flac', '.wav', '.m4a', '.aac', '.ogg', '.wma'],
    'VIDEO': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm'],
    'ARQUIVOS': ['.zip', '.rar', '.7z', '.tar', '.gz']
}

HASH_ALGORITHMS = ['md5', 'sha1', 'sha256', 'sha512']

class ForensicAnalyzer:
    """Analisador forense completo para metadados de arquivos"""
    
    def __init__(self):
        self.analyzers = {
            'IMAGENS': self.analyze_image,
            'PDF': self.analyze_pdf,
            'DOCUMENTOS': self.analyze_document,
            'AUDIO': self.analyze_audio,
            'VIDEO': self.analyze_video,
            'ARQUIVOS': self.analyze_archive
        }
        try:
            self.mime_detector = magic.Magic(mime=True)
        except Exception:
            self.mime_detector = None
            logging.warning("Magic não disponível - usando apenas extensões")

    def analyze_image(self, file_path: Path) -> Dict[str, Any]:
        """Análise avançada de imagens"""
        try:
            with Image.open(file_path) as img:
                info = {
                    "Tipo": "Imagem",
                    "Formato": img.format,
                    "Tamanho": f"{img.size[0]}x{img.size[1]}",
                    "Modo": img.mode,
                    "EXIF": {}
                }
                
                try:
                    exif_data = img._getexif()
                    if exif_data:
                        for tag_id, value in exif_data.items():
                            tag_name = ExifTags.TAGS.get(tag_id, tag_id)
                            try:
                                # Tentar serializar o valor
                                if isinstance(value, (str, int, float, bool)):
                                    info["EXIF"][tag_name] = value
                                else:
                                    info["EXIF"][tag_name] = str(value)
                            except (TypeError, ValueError):
                                info["EXIF"][tag_name] = str(value)
                except Exception as e:
                    info["EXIF_Erro"] = str(e)
                    
                return info
        except Exception as e:
            logging.error(f"Erro análise imagem {file_path}: {e}")
            return {"Tipo": "Imagem", "Erro": str(e)}

    def analyze_pdf(self, file_path: Path) -> Dict[str, Any]:
        """Análise de PDF"""
        try:
            reader = PdfReader(file_path)
            info = {
                "Tipo": "PDF",
                "Páginas": len(reader.pages),
                "Criptografado": reader.is_encrypted,
                "Metadados": {}
            }
            
            if reader.metadata:
                for key, value in reader.metadata.items():
                    info["Metadados"][key.replace('/', '')] = str(value)
            
            return info
        except Exception as e:
            logging.error(f"Erro análise PDF {file_path}: {e}")
            return {"Tipo": "PDF", "Erro": str(e)}

    def analyze_document(self, file_path: Path) -> Dict[str, Any]:
        """Análise de documentos Office"""
        ext = file_path.suffix.lower()
        try:
            if ext == '.docx':
                doc = docx.Document(file_path)
                props = doc.core_properties
                return {
                    "Tipo": "Documento Word",
                    "Formato": "DOCX",
                    "Autor": props.author or "N/A",
                    "Criado": str(props.created) if props.created else "N/A",
                    "Modificado": str(props.modified) if props.modified else "N/A",
                    "Parágrafos": len(doc.paragraphs)
                }
            elif ext == '.xlsx':
                wb = openpyxl.load_workbook(file_path, read_only=True)
                return {
                    "Tipo": "Planilha Excel",
                    "Formato": "XLSX",
                    "Abas": wb.sheetnames,
                    "Quantidade_Abas": len(wb.sheetnames)
                }
            elif ext == '.pptx':
                prs = pptx.Presentation(file_path)
                return {
                    "Tipo": "Apresentação PowerPoint",
                    "Formato": "PPTX",
                    "Slides": len(prs.slides)
                }
            elif ext == '.txt':
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read()
                    return {
                        "Tipo": "Documento Texto",
                        "Linhas": len(content.splitlines()),
                        "Caracteres": len(content)
                    }
            else:
                return {"Tipo": f"Documento {ext}", "Formato": ext.upper()}
        except Exception as e:
            logging.error(f"Erro análise documento {file_path}: {e}")
            return {"Tipo": f"Documento {ext}", "Erro": str(e)}

    def analyze_audio(self, file_path: Path) -> Dict[str, Any]:
        """Análise de áudio"""
        try:
            audio = mutagen.File(file_path)
            if audio:
                info = {
                    "Tipo": "Áudio",
                    "Duração": round(audio.info.length, 2),
                    "Bitrate": getattr(audio.info, 'bitrate', 'N/A'),
                    "Sample_Rate": getattr(audio.info, 'sample_rate', 'N/A'),
                    "Metadados": {}
                }
                
                if audio.tags:
                    for key, value in audio.tags.items():
                        try:
                            info["Metadados"][str(key)] = str(value)[:100]  # Limitar tamanho
                        except:
                            pass
                
                return info
            return {"Tipo": "Áudio", "Erro": "Metadados não disponíveis"}
        except Exception as e:
            logging.error(f"Erro análise áudio {file_path}: {e}")
            return {"Tipo": "Áudio", "Erro": str(e)}

    def analyze_video(self, file_path: Path) -> Dict[str, Any]:
        """Análise de vídeo"""
        try:
            cap = cv2.VideoCapture(str(file_path))
            if not cap.isOpened():
                return {"Tipo": "Vídeo", "Erro": "Não foi possível abrir"}
            
            frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            duration = frames / fps if fps > 0 else 0
            
            cap.release()
            
            return {
                "Tipo": "Vídeo",
                "Resolução": f"{width}x{height}",
                "FPS": round(fps, 2),
                "Duração(s)": round(duration, 2),
                "Frames": frames
            }
        except Exception as e:
            logging.error(f"Erro análise vídeo {file_path}: {e}")
            return {"Tipo": "Vídeo", "Erro": str(e)}

    def analyze_archive(self, file_path: Path) -> Dict[str, Any]:
        """Análise de arquivos compactados"""
        try:
            if zipfile.is_zipfile(file_path):
                with zipfile.ZipFile(file_path, 'r') as z:
                    file_list = z.namelist()
                    return {
                        "Tipo": "Arquivo ZIP",
                        "Total_Arquivos": len(file_list),
                        "Arquivos": file_list[:10]  # Primeiros 10 arquivos
                    }
            return {"Tipo": "Arquivo compactado", "Formato": file_path.suffix}
        except Exception as e:
            logging.error(f"Erro análise arquivo {file_path}: {e}")
            return {"Tipo": "Arquivo compactado", "Erro": str(e)}

    def analyze_file(self, file_path: Path) -> Dict[str, Any]:
        """Analisa arquivo baseado na extensão"""
        ext = file_path.suffix.lower()
        
        for category, extensions in SUPPORTED_EXTENSIONS.items():
            if ext in extensions:
                analyzer = self.analyzers.get(category)
                if analyzer:
                    return analyzer(file_path)
        
        # Fallback para detecção MIME
        if self.mime_detector:
            try:
                mime_type = self.mime_detector.from_file(str(file_path))
                if mime_type.startswith('image/'):
                    return self.analyze_image(file_path)
                elif mime_type.startswith('audio/'):
                    return self.analyze_audio(file_path)
                elif mime_type.startswith('video/'):
                    return self.analyze_video(file_path)
                elif mime_type == 'application/pdf':
                    return self.analyze_pdf(file_path)
            except Exception:
                pass
        
        return {"Tipo": "Desconhecido", "Extensão": ext}

def calculate_hashes(file_path: Path) -> Dict[str, str]:
    """Calcula hashes do arquivo"""
    hashes = {}
    try:
        file_size = file_path.stat().st_size
        if file_size > 100 * 1024 * 1024:  # 100MB limite
            return {"AVISO": "Arquivo muito grande para cálculo de hash"}
        
        with open(file_path, 'rb') as f:
            chunk_size = 8192
            hashers = {algo: hashlib.new(algo) for algo in HASH_ALGORITHMS}
            
            while chunk := f.read(chunk_size):
                for hasher in hashers.values():
                    hasher.update(chunk)
            
            for algo, hasher in hashers.items():
                hashes[algo.upper()] = hasher.hexdigest()
                
        return hashes
    except Exception as e:
        logging.error(f"Erro ao calcular hash para {file_path}: {e}")
        return {"ERRO": str(e)}

def process_single_file(file_path: Path, analyzer: ForensicAnalyzer) -> Dict[str, Any]:
    """Processa um único arquivo"""
    try:
        file_stat = file_path.stat()
        
        metadata = {
            "Nome_Arquivo": file_path.name,
            "Caminho_Absoluto": str(file_path.absolute()),
            "Tamanho_Bytes": file_stat.st_size,
            "Data_Modificacao": datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
            "Extensao": file_path.suffix.lower(),
            "Analise_Concluida": False
        }
        
        # Análise específica do tipo de arquivo
        file_analysis = analyzer.analyze_file(file_path)
        metadata.update(file_analysis)
        
        # Calcular hashes (apenas para arquivos menores)
        if file_stat.st_size <= 50 * 1024 * 1024:  # 50MB
            metadata["Hashes"] = calculate_hashes(file_path)
        
        metadata["Analise_Concluida"] = True
        return metadata
        
    except Exception as e:
        logging.error(f"Erro processando {file_path}: {e}")
        return {
            "Nome_Arquivo": file_path.name,
            "Caminho_Absoluto": str(file_path.absolute()),
            "Erro_Critico": str(e),
            "Analise_Concluida": False
        }

class AnalysisManager:
    """Gerenciador de análises"""
    
    def __init__(self):
        self.analyzer = ForensicAnalyzer()
        from database import ResultsDatabase
        self.db = ResultsDatabase()
        self.active_analyses = {}
        self._lock = threading.Lock()

    def count_supported_files(self, directory: Path) -> int:
        """Conta arquivos suportados"""
        count = 0
        try:
            supported_extensions = set(
                ext for sublist in SUPPORTED_EXTENSIONS.values() for ext in sublist
            )
            for file_path in directory.rglob('*'):
                if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                    count += 1
                    if count > 100000:  # Limite de segurança
                        break
        except Exception as e:
            logging.error(f"Erro ao contar arquivos: {e}")
        return count

    def stream_supported_files(self, directory: Path, max_files: int = 50000) -> Generator[Path, None, None]:
        """Gera arquivos suportados"""
        count = 0
        try:
            supported_extensions = set(
                ext for sublist in SUPPORTED_EXTENSIONS.values() for ext in sublist
            )
            for file_path in directory.rglob('*'):
                if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                    yield file_path
                    count += 1
                    if count >= max_files:
                        return
        except Exception as e:
            logging.error(f"Erro ao coletar arquivos: {e}")

    def start_analysis(self, analysis_id: str, directory_path: str) -> bool:
        """Inicia uma análise"""
        try:
            path_obj = Path(directory_path)
            if not path_obj.exists() or not path_obj.is_dir():
                return False
            
            total_files = self.count_supported_files(path_obj)
            
            with self._lock:
                self.active_analyses[analysis_id] = {
                    "directory": directory_path,
                    "total_files": total_files,
                    "start_time": datetime.now(),
                    "status": "running"
                }
            
            # Salvar metadados
            self.db.save_analysis_metadata(analysis_id, directory_path, total_files)
            
            # Executar em thread
            thread = threading.Thread(
                target=self._run_analysis,
                args=(analysis_id, path_obj),
                daemon=True
            )
            thread.start()
            
            return True
            
        except Exception as e:
            logging.error(f"Erro ao iniciar análise: {e}")
            return False

    def _run_analysis(self, analysis_id: str, directory: Path):
        """Executa a análise"""
        try:
            with ThreadPoolExecutor(max_workers=4) as executor:
                future_to_file = {}
                
                for file_path in self.stream_supported_files(directory):
                    future = executor.submit(process_single_file, file_path, self.analyzer)
                    future_to_file[future] = file_path
                
                for future in as_completed(future_to_file):
                    try:
                        result = future.result()
                        self.db.save_result(analysis_id, result)
                    except Exception as e:
                        logging.error(f"Erro no processamento: {e}")
            
            self.db.update_analysis_status(analysis_id, "completed")
            
            with self._lock:
                if analysis_id in self.active_analyses:
                    self.active_analyses[analysis_id]["status"] = "completed"
                    self.active_analyses[analysis_id]["end_time"] = datetime.now()
            
            logging.info(f"Análise {analysis_id} concluída")
            
        except Exception as e:
            logging.error(f"Erro na análise {analysis_id}: {e}")
            self.db.update_analysis_status(analysis_id, "error", str(e))

    def get_analysis_status(self, analysis_id: str) -> Optional[Dict[str, Any]]:
        """Obtém status da análise"""
        return self.db.get_analysis_status(analysis_id)

    def get_analysis_results(self, analysis_id: str, limit: int = 1000, offset: int = 0) -> List[Dict[str, Any]]:
        """Obtém resultados da análise"""
        return self.db.get_analysis_results(analysis_id, limit, offset)

    def get_analysis_stats(self, analysis_id: str) -> Dict[str, Any]:
        """Obtém estatísticas da análise"""
        return self.db.get_analysis_stats(analysis_id)